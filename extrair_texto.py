import pdfplumber
import os
from openpyxl import load_workbook
from tkinter import messagebox, filedialog
import tkinter as tk
import sys

def ler_pdf(pdf_path):   

    pdf = pdfplumber.open(pdf_path)
    paginas = pdf.pages
    pagina = paginas[0]
    texto = pagina.extract_text()
    return texto

def extrair_resultado(texto):
    info = {
            "Contrato": None,  #OK
            "COD": [], #OK
            "Desc": [], #OK
            "QTD": [], #OK
            "QTD_Total": None, #OK
            "Cliente": None,  #OK
            "Porto": None,  #OK
            "Destino": None, #OK
            "USD": [], #OK
            "Val": [], #OK
            }
    
    try:
        info["Contrato"] = texto.split("INSTRUÇÃO FÁBRICA - ")[1].split("Cliente")[0].strip()
        info["Cliente"] = texto.split("Cliente")[1].split("ADDRESS")[0].strip()
        info["Porto"] = texto.split("Destino")[1].split("Enbarque")[0].strip()
        info["Destino"] = texto.split("Mercado")[1].split("Temperatura")[0].strip()
        info["QTD_Total"] = float(texto.split("Total ")[1].split("US$")[0].strip().replace(",","."))*1000

        produtos = texto.split("Expire Date")[1].split("External Label")[0].strip()
    
        linhas = (produtos.split("\n"))
        for linha in linhas:
            if "CIF" in linha:
                break
            else:
                info["COD"].append(int(linha.split()[0].strip()))
                info["Desc"].append(linha.split(" CARTONS ")[0].split(" ")[1].strip())
                info["QTD"].append(float((linha.split(" CARTONS ")[1].split(" ")[0].strip()).replace(",","."))*1000)
                info["USD"].append(float(linha.split(" CARTONS ")[1].split(" ")[1].strip().replace(".","").replace(",","."))/1000)
                n = len(linha.split(" "))
                info["Val"].append(linha.split(" ")[n-2]+ " " + linha.split(" ")[n-1])
    except Exception as e:
        messagebox.showerror("Erro", f"Erro ao extrair dados do PDF: {e}")
        sys.exit(1)
    
    return info

def registrar_resultados(resultados, caminho_excel):

    if not caminho_excel.endswith(".xlsx"):
        messagebox.showerror("O arquivo descrito em config.txt não é um excel. Verificar")
        sys.exit(1)
    
    wb = load_workbook(caminho_excel)
    if "RESULTADOS" not in wb.sheetnames:
        messagebox.showerror("Erro", "A aba 'RESULTADOS' não existe no Excel.")
        sys.exit(1)
    
    ws = wb["RESULTADOS"]

    linha = 1
    while ws.cell(row=linha, column=1).value is not None:
        linha += 1
    
    tamanhos = [len(resultados["COD"]), 
                len(resultados["Desc"]),
                len(resultados["QTD"]), 
                len(resultados["USD"]), 
                len(resultados["Val"])
    ]

    if len(set(tamanhos)) >1:    
        messagebox.showerror("Erro de extração, o numero de informações não é valido")
        sys.exit(1)
    else:
        t = tamanhos[0]

    listas = [[resultados["Contrato"]]*t, 
              resultados["COD"], 
              resultados["Desc"], 
              resultados["QTD"], 
              [resultados["QTD_Total"]]*t,
              [resultados["Cliente"]]*t,
              [resultados["Porto"]]*t,
              [resultados["Destino"]]*t,
              resultados["USD"], 
              resultados["Val"]
    ]
    
    for x in range(t):
        for n, lista in enumerate(listas):
            ws.cell(row=linha+x, column=n+1, value=lista[x])

    wb.save(caminho_excel)
    #print(listas)
        
def carregar_caminho_excel(caminho_arquivo_txt):
    try:
        with open(caminho_arquivo_txt, 'r', encoding='utf-8') as f:
            caminho = f.readline().strip()
            return caminho
    except FileNotFoundError:
        messagebox.showerror("Erro", f"Arquivo {caminho_arquivo_txt} não encontrado. Verificar diretório do executável")
        sys.exit(1)


def main():
    root = tk.Tk()
    root.withdraw()

    pasta_pdf = filedialog.askdirectory(title="Selecione a pasta com os PDF's")

    if not pasta_pdf:
        messagebox.showerror("Erro", "Nenhuma pasta foi selecionada")
        sys.exit(1)
    
    arquivos = os.listdir(pasta_pdf)
    pdfs = []
    caminho_excel = carregar_caminho_excel("config.txt")

    for arquivo in arquivos:
        if arquivo.endswith(".pdf"):    
            pdfs.append(arquivo)

    if len(pdfs) < 1:
        messagebox.showerror("Erro", "Não foram encontrados arquivos PDF na pasta selecionada")
        sys.exit(1)
    
    for arquivo in arquivos:
        try:
            if arquivo.endswith(".pdf"): 
                texto = ler_pdf(arquivo)
                resultados = extrair_resultado(texto)
                registrar_resultados(resultados, caminho_excel)
                print(f"Arquivo {arquivo} extraido com sucesso")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro inesperado {e}")
        
    messagebox.showinfo("Finalizando", "Finalizando programa")

    
if __name__ == "__main__":
    main()