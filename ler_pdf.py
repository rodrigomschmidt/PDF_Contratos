import pdfplumber
import os
from openpyxl import load_workbook
from tkinter import filedialog, messagebox
import tkinter as tk

def ler_pdf(pasta_pdf):
    tipos = []
    textos = []
    arquivos = os.listdir(pasta_pdf)
    
    for arquivo in arquivos:
        pdf_path = os.path.join(pasta_pdf,arquivo)
        with pdfplumber.open(pdf_path) as pdf:
            pagina = pdf.pages[0]
            texto = pagina.extract_text()

        #print(texto)
        textos.append(texto)
        if "Avermectina" in texto:
            tipos.append("A")
            #print("Laudo de Avermectina")
        else:
            tipos.append("M")
            #print("Laudo microbiológico")

    return tipos, textos


def extrair_resultados(tipos, textos):
    # Avermectina - Puxar numero do laudo e data de coleta
    # Micro - Puxar numero do laudo e lote
    laudos =[]
    infos = []
    
    for n, texto in enumerate(textos):
        
        print(f"Texto n: {n} do tipo {tipos[n]}")
        laudo = texto.split("Solicitante")[0].split("Relatório de ensaio n°: ")[1]
        laudo = laudo.strip()
        print(laudo)
        laudos.append(laudo)
        
        if tipos[n] == "M": #puxar o lote
            divisor_1 = "Data de Abate"
            divisor_2 = "Lote: "
            info = texto.split(divisor_1)[0].split(divisor_2)[1]
            info = info.strip()
            print(info)
            infos.append(info)
        else:
            divisor_1 = "Local da coleta"
            divisor_2 = "Data de Abate: "
            info = texto.split(divisor_1)[0].split(divisor_2)[1]
            info = info.strip()
            print(info)
            infos.append(info)

    return laudos, infos
        

def inserir_excel(caminho_excel, laudos, infos, tipos, root):
    
    wb = load_workbook(caminho_excel)
        
    if not (len(tipos) == len(laudos) == len(infos)):
        messagebox.showerror("Erro", "Numero de informações imcompatíveis")
        return False
    
    for n, _ in enumerate(laudos):
        resultados = [infos[n], laudos[n]]

        if tipos[n] == "M":
            aba_destino = "Microbiológico"
        else:
            aba_destino = "Avermectina"

        ws = wb[aba_destino]

        linha = 1
        while ws.cell(row=linha, column=1).value is not None:
            linha += 1

        for i, valor in enumerate(resultados, start=1):  # coluna A, B, C...
            ws.cell(row=linha, column=i, value=valor)
        
    try:
        wb.save(caminho_excel)
        return True
    except:
        messagebox.showerror("Erro", "Não foi possivel salvar a planilha. Verifique se o arquivo está aberto em algum local")
        return False
    
def carregar_caminho_excel(caminho_arquivo_txt):
    try:
        with open(caminho_arquivo_txt, 'r', encoding='utf-8') as f:
            caminho = f.readline().strip()
            return caminho
    except FileNotFoundError:
        messagebox.showerror("Erro", f"Arquivo {caminho_arquivo_txt} não encontrado.")
        return None

def main():
    root = tk.Tk()
    root.withdraw()
    
    pasta_pdf = filedialog.askdirectory(title="Selecione a pasta com os laudos")

    if not pasta_pdf:
        messagebox.showerror("Erro", "Nenhuma pasta foi selecionada")
        return
    
    tipos, textos = ler_pdf(pasta_pdf)

    caminho_excel = carregar_caminho_excel("caminho_excel.txt")
    if not caminho_excel:
        messagebox.showerror("Erro", "Verificar arquivo de caminho do excel")
        return
    
    laudos, infos = extrair_resultados(tipos, textos)
    check = inserir_excel(caminho_excel, laudos, infos, tipos, root)

    if check == True:
        messagebox.showinfo("Concluído", "Dados inseridos com sucesso")
        root.destroy()
    else:
        messagebox.showinfo("Fechando", "Fechando aplicação")
        root.destroy()

if __name__ == "__main__":

    main()
    
    